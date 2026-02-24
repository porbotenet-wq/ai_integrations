"""Excel import/export routes"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession
from bot.db.session import async_session
from bot.db.models import (
    Crew, WorkType, FloorVolume, GPRWeekly, DailyProgress,
    DailyPlanFact, ConstructionObject,
)
from datetime import datetime
import io

router = APIRouter(prefix="/api/excel", tags=["excel"])


async def get_db():
    async with async_session() as session:
        yield session


@router.post("/import/{object_id}")
async def import_excel(
    object_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Импорт Excel файла план-факт в БД"""
    import openpyxl

    obj = await db.get(ConstructionObject, object_id)
    if not obj:
        raise HTTPException(404, "Object not found")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    stats = {"crews": 0, "work_types": 0, "floor_volumes": 0, "gpr_weekly": 0, "daily_progress": 0, "plan_fact": 0}

    # --- Бригады ---
    if '👷 Бригады' in wb.sheetnames:
        ws = wb['👷 Бригады']
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            code = row[0]
            if not code or str(code).startswith('ID'):
                continue
            name = str(row[1]) if row[1] else str(code)
            existing = (await db.execute(select(Crew).where(Crew.code == str(code)))).scalar_one_or_none()
            if not existing:
                crew = Crew(
                    code=str(code), name=name,
                    foreman=str(row[2]) if row[2] and str(row[2]) != '—' else None,
                    phone=str(row[3]) if row[3] and str(row[3]) != '—' else None,
                    specialization=str(row[4]) if row[4] else None,
                    max_workers=int(row[5]) if row[5] and str(row[5]).isdigit() else 0,
                    status='standby' if row[6] and 'Ожид' in str(row[6]) else 'active',
                    object_id=object_id,
                )
                db.add(crew)
                stats["crews"] += 1
        await db.flush()

    # --- Виды работ ---
    if '📝 Виды работ' in wb.sheetnames:
        ws = wb['📝 Виды работ']
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            code = row[1]
            if not code or str(code).startswith('Код'):
                continue
            existing = (await db.execute(select(WorkType).where(WorkType.code == str(code)))).scalar_one_or_none()
            if not existing:
                crew_code = str(row[7]) if row[7] and str(row[7]) != '—' else None
                crew = None
                if crew_code:
                    crew = (await db.execute(select(Crew).where(Crew.code == crew_code))).scalar_one_or_none()
                wt = WorkType(
                    code=str(code), name=str(row[2]),
                    unit=str(row[3]) if row[3] else 'шт',
                    category=str(row[4]) if row[4] else None,
                    sequence_order=int(row[5]) if row[5] else 0,
                    requires_inspection=str(row[6]).strip() == 'Да' if row[6] else False,
                    default_crew_id=crew.id if crew else None,
                )
                db.add(wt)
                stats["work_types"] += 1
        await db.flush()

    # Build maps
    crews_r = await db.execute(select(Crew))
    crews_map = {c.code: c.id for c in crews_r.scalars().all()}
    wt_r = await db.execute(select(WorkType))
    wt_map = {w.code: w.id for w in wt_r.scalars().all()}

    # --- Объёмы по этажам ---
    if '🏗️ Объёмы по этажам' in wb.sheetnames:
        ws = wb['🏗️ Объёмы по этажам']
        col_wt = [(2, 3, 'МОД'), (5, 6, 'БУР12'), (8, 9, 'БУР16'),
                  (11, 12, 'КРН-Н'), (14, 15, 'КРН-В'), (17, 18, 'РЕЗ'), (20, 21, 'ГЕР')]
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            if not row[0] or not str(row[0]).isdigit():
                continue
            floor_val = int(row[0])
            facade = str(row[1])
            for pc, fc, wc in col_wt:
                wt_id = wt_map.get(wc)
                if not wt_id:
                    continue
                plan = float(row[pc]) if row[pc] else 0
                fact = float(row[fc]) if row[fc] else 0
                if plan == 0 and fact == 0:
                    continue
                await db.execute(text("""
                    INSERT INTO floor_volumes (object_id, floor, facade, work_type_id, plan_qty, fact_qty)
                    VALUES (:oid, :fl, :fa, :wt, :p, :f)
                    ON CONFLICT (object_id, floor, facade, work_type_id)
                    DO UPDATE SET plan_qty = :p, fact_qty = :f
                """), {"oid": object_id, "fl": floor_val, "fa": facade, "wt": wt_id, "p": plan, "f": fact})
                stats["floor_volumes"] += 1

    # --- План-Факт ---
    if '📋 План-Факт' in wb.sheetnames:
        ws = wb['📋 План-Факт']
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            date_val = row[0]
            if not date_val:
                continue
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, str) and '.' in date_val:
                try:
                    date_str = datetime.strptime(date_val.strip(), '%d.%m.%Y').strftime('%Y-%m-%d')
                except:
                    continue
            else:
                continue

            wc = str(row[5]) if row[5] else None
            cc = str(row[11]) if row[11] and str(row[11]) != '—' else None

            pf = DailyPlanFact(
                object_id=object_id, date=date_str,
                work_name=str(row[4]) if row[4] else None,
                work_code=wc,
                sequence_order=int(row[6]) if row[6] and str(row[6]).isdigit() else None,
                plan_daily=float(row[7]) if row[7] else None,
                fact_volume=float(row[8]) if row[8] else None,
                crew_code=cc,
                workers_count=int(row[12]) if row[12] and str(row[12]).isdigit() else None,
                inspection_status=str(row[14]) if row[14] else 'Нет',
                day_number=int(row[1]) if row[1] and str(row[1]).isdigit() else None,
                floor=int(row[2]) if row[2] and str(row[2]).isdigit() else None,
                facade=str(row[3]) if row[3] else None,
                work_type_id=wt_map.get(wc),
                crew_id=crews_map.get(cc) if cc else None,
                unit='шт',
            )
            db.add(pf)
            stats["plan_fact"] += 1

    await db.commit()
    return {"status": "ok", "object_id": object_id, "imported": stats}


@router.get("/export/{object_id}")
async def export_excel(object_id: int, db: AsyncSession = Depends(get_db)):
    """Экспорт данных объекта в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    obj = await db.get(ConstructionObject, object_id)
    if not obj:
        raise HTTPException(404, "Object not found")

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font_w = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # --- Sheet 1: Дашборд ---
    ws = wb.active
    ws.title = "📊 Дашборд"
    ws.append([f"ДАШБОРД — {obj.name}"])
    ws.append(["Адрес", obj.address or ""])
    ws.append(["Период", f"{obj.contract_date} — {obj.deadline_date}"])
    ws.append([])

    kpi_r = await db.execute(text("""
        SELECT wt.name, wt.unit,
            COALESCE(SUM(fv.plan_qty),0), COALESCE(SUM(fv.fact_qty),0)
        FROM floor_volumes fv JOIN work_types wt ON wt.id = fv.work_type_id
        WHERE fv.object_id = :oid
        GROUP BY wt.name, wt.unit, wt.sequence_order ORDER BY wt.sequence_order
    """), {"oid": object_id})
    ws.append(["Показатель", "Ед.изм.", "План", "Факт", "% выполнения"])
    for r in kpi_r:
        plan, fact = float(r[2]), float(r[3])
        pct = round(fact / plan * 100, 1) if plan > 0 else 0
        ws.append([r[0], r[1], plan, fact, f"{pct}%"])

    # --- Sheet 2: Бригады ---
    ws2 = wb.create_sheet("👷 Бригады")
    ws2.append(["Код", "Название", "Бригадир", "Телефон", "Специализация", "Макс. людей", "Статус"])
    crews_r = await db.execute(select(Crew).order_by(Crew.code))
    for c in crews_r.scalars().all():
        ws2.append([c.code, c.name, c.foreman or "—", c.phone or "—",
                     c.specialization, c.max_workers, c.status])

    # --- Sheet 3: Виды работ ---
    ws3 = wb.create_sheet("📝 Виды работ")
    ws3.append(["Код", "Вид работы", "Ед.изм.", "Категория", "Порядок", "Требует ТН"])
    wt_r = await db.execute(select(WorkType).order_by(WorkType.sequence_order))
    for w in wt_r.scalars().all():
        ws3.append([w.code, w.name, w.unit, w.category, w.sequence_order,
                     "Да" if w.requires_inspection else "Нет"])

    # --- Sheet 4: План-Факт ---
    ws4 = wb.create_sheet("📋 План-Факт")
    ws4.append(["Дата", "День", "Этаж", "Фасад", "Вид работы", "Код",
                "План", "Факт", "Отклонение", "%", "Бригада", "Людей",
                "Приёмка ТН", "Примечание"])
    pf_r = await db.execute(
        select(DailyPlanFact)
        .where(DailyPlanFact.object_id == object_id)
        .order_by(DailyPlanFact.date, DailyPlanFact.sequence_order)
    )
    for r in pf_r.scalars().all():
        ws4.append([
            r.date.strftime('%d.%m.%Y') if r.date else "",
            r.day_number, r.floor, r.facade, r.work_name, r.work_code,
            r.plan_daily, r.fact_volume, r.deviation, r.completion_pct,
            r.crew_code, r.workers_count, r.inspection_status, r.notes,
        ])

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"SPK_{obj.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
